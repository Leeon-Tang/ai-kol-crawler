# -*- coding: utf-8 -*-
"""
GitHub开发者导出任务
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from backend.utils.logger import setup_logger
from backend.utils.config_loader import load_config
from backend.storage.repositories.github_repository import GitHubRepository

logger = setup_logger()


class GitHubExportTask:
    """GitHub开发者导出任务"""
    
    def __init__(self, repository: GitHubRepository):
        self.repository = repository
        self.config = load_config()
    
    def run(self, limit: int = 1000) -> str:
        """
        导出GitHub开发者数据到Excel
        
        Args:
            limit: 最大导出数量
            
        Returns:
            导出文件路径
        """
        logger.info("开始导出GitHub开发者数据")
        
        # 获取合格的开发者
        developers = self.repository.get_qualified_developers(limit=limit)
        
        if not developers:
            logger.warning("没有可导出的开发者数据")
            return ""
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "GitHub Developers"
        
        # 设置表头
        headers = [
            "用户名", "姓名", "个人主页", "简介", "公司", "位置",
            "博客", "Twitter", "Email", "联系方式",
            "公开仓库数", "Followers", "Following",
            "分析仓库数", "总Stars", "总Forks", "平均Stars", "平均Forks",
            "主要语言", "原创仓库数", "发现时间"
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row, dev in enumerate(developers, 2):
            import json
            top_languages = json.loads(dev.get('top_languages', '[]'))
            top_languages_str = ', '.join(top_languages) if top_languages else ''
            
            data = [
                dev.get('username', ''),
                dev.get('name', ''),
                dev.get('profile_url', ''),
                dev.get('bio', ''),
                dev.get('company', ''),
                dev.get('location', ''),
                dev.get('blog', ''),
                dev.get('twitter', ''),
                dev.get('email', ''),
                dev.get('contact_info', ''),
                dev.get('public_repos', 0),
                dev.get('followers', 0),
                dev.get('following', 0),
                dev.get('analyzed_repos', 0),
                dev.get('total_stars', 0),
                dev.get('total_forks', 0),
                dev.get('avg_stars', 0),
                dev.get('avg_forks', 0),
                top_languages_str,
                dev.get('original_repos', 0),
                dev.get('discovered_at', '')
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # 保存文件
        export_dir = self.config.get('export', {}).get('output_dir', 'exports')
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"github_developers_{timestamp}.xlsx"
        filepath = os.path.join(export_dir, filename)
        
        wb.save(filepath)
        logger.info(f"导出完成: {filepath}")
        logger.info(f"共导出 {len(developers)} 个开发者")
        
        return filepath
    
    def run_today(self, today_start, today_end) -> str:
        """
        导出今日GitHub开发者数据到Excel
        
        Args:
            today_start: 今天开始时间
            today_end: 今天结束时间
            
        Returns:
            导出文件路径
        """
        from datetime import timedelta, timezone
        logger.info("开始导出今日GitHub开发者数据")
        
        # 格式化时间为字符串
        start_str = today_start.strftime('%Y-%m-%d %H:%M:%S')
        end_str = today_end.strftime('%Y-%m-%d %H:%M:%S')
        
        # 查询今日数据
        query = """
            SELECT * FROM github_developers 
            WHERE status = 'qualified' AND is_indie_developer = 1
            AND discovered_at >= ? AND discovered_at <= ?
            ORDER BY total_stars DESC, followers DESC
        """
        developers = self.repository.db.fetchall(query, (start_str, end_str))
        
        if not developers:
            logger.warning("今天没有可导出的开发者数据")
            return ""
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "GitHub Developers Today"
        
        # 设置表头
        headers = [
            "用户名", "姓名", "个人主页", "简介", "公司", "位置",
            "博客", "Twitter", "Email", "联系方式",
            "公开仓库数", "Followers", "Following",
            "分析仓库数", "总Stars", "总Forks", "平均Stars", "平均Forks",
            "主要语言", "原创仓库数", "发现时间"
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row, dev in enumerate(developers, 2):
            import json
            top_languages = json.loads(dev.get('top_languages', '[]'))
            top_languages_str = ', '.join(top_languages) if top_languages else ''
            
            data = [
                dev.get('username', ''),
                dev.get('name', ''),
                dev.get('profile_url', ''),
                dev.get('bio', ''),
                dev.get('company', ''),
                dev.get('location', ''),
                dev.get('blog', ''),
                dev.get('twitter', ''),
                dev.get('email', ''),
                dev.get('contact_info', ''),
                dev.get('public_repos', 0),
                dev.get('followers', 0),
                dev.get('following', 0),
                dev.get('analyzed_repos', 0),
                dev.get('total_stars', 0),
                dev.get('total_forks', 0),
                dev.get('avg_stars', 0),
                dev.get('avg_forks', 0),
                top_languages_str,
                dev.get('original_repos', 0),
                dev.get('discovered_at', '')
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # 保存文件
        export_dir = self.config.get('export', {}).get('output_dir', 'exports')
        os.makedirs(export_dir, exist_ok=True)
        
        beijing_time = datetime.now(timezone.utc) + timedelta(hours=8)
        today_date = beijing_time.strftime('%Y%m%d')
        timestamp = beijing_time.strftime('%Y%m%d_%H%M%S')
        filename = f"github_developers_today_{timestamp}.xlsx"
        filepath = os.path.join(export_dir, filename)
        
        wb.save(filepath)
        logger.info(f"导出今日数据完成: {filepath}")
        logger.info(f"共导出 {len(developers)} 个开发者")
        
        return filepath
