# -*- coding: utf-8 -*-
"""
GitHub开发者导出任务
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from utils.logger import setup_logger
from utils.config_loader import load_config
from storage.repositories.github_repository import GitHubRepository

logger = setup_logger()


class GitHubExportTask:
    """GitHub开发者导出任务"""
    
    def __init__(self, repository: GitHubRepository):
        self.repository = repository
        self.config = load_config()
    
    def run(self, limit: int = 1000) -> str:
        """
        导出GitHub开发者数据到Excel
        
        Args:
            limit: 最大导出数量
            
        Returns:
            导出文件路径
        """
        logger.info("开始导出GitHub开发者数据")
        
        # 获取合格的开发者
        developers = self.repository.get_qualified_developers(limit=limit)
        
        if not developers:
            logger.warning("没有可导出的开发者数据")
            return ""
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "GitHub Developers"
        
        # 设置表头
        headers = [
            "用户名", "姓名", "个人主页", "简介", "公司", "位置",
            "博客", "Twitter", "Email", "联系方式",
            "公开仓库数", "Followers", "Following",
            "分析仓库数", "总Stars", "总Forks", "平均Stars", "平均Forks",
            "主要语言", "原创仓库数", "发现时间"
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row, dev in enumerate(developers, 2):
            import json
            top_languages = json.loads(dev.get('top_languages', '[]'))
            top_languages_str = ', '.join(top_languages) if top_languages else ''
            
            data = [
                dev.get('username', ''),
                dev.get('name', ''),
                dev.get('profile_url', ''),
                dev.get('bio', ''),
                dev.get('company', ''),
                dev.get('location', ''),
                dev.get('blog', ''),
                dev.get('twitter', ''),
                dev.get('email', ''),
                dev.get('contact_info', ''),
                dev.get('public_repos', 0),
                dev.get('followers', 0),
                dev.get('following', 0),
                dev.get('analyzed_repos', 0),
                dev.get('total_stars', 0),
                dev.get('total_forks', 0),
                dev.get('avg_stars', 0),
                dev.get('avg_forks', 0),
                top_languages_str,
                dev.get('original_repos', 0),
                dev.get('discovered_at', '')
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # 保存文件
        export_dir = self.config.get('export', {}).get('output_dir', 'exports')
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"github_developers_{timestamp}.xlsx"
        filepath = os.path.join(export_dir, filename)
        
        wb.save(filepath)
        logger.info(f"导出完成: {filepath}")
        logger.info(f"共导出 {len(developers)} 个开发者")
        
        return filepath
