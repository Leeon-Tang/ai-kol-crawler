"""
导出任务 - 导出KOL列表到Excel
"""
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from utils.logger import setup_logger


logger = setup_logger()


class ExportTask:
    """导出任务"""
    
    def __init__(self, repository, config_path='config/config.json'):
        self.repository = repository
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        self.export_config = config['export']
    
    def run(self):
        """
        执行导出任务
        导出合格KOL到Excel
        """
        logger.info("=" * 50)
        logger.info("开始执行导出任务")
        logger.info("=" * 50)
        
        # 获取所有合格的KOL
        kols = self.repository.get_qualified_kols()
        
        if not kols:
            logger.info("没有合格的KOL可导出")
            return
        
        logger.info(f"待导出KOL数: {len(kols)}")
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "AI KOL列表"
        
        # 设置表头
        headers = [
            "频道名称",
            "频道链接",
            "订阅数",
            "AI内容占比",
            "平均观看数",
            "平均点赞数",
            "平均评论数",
            "互动率",
            "总视频数",
            "最后视频日期",
            "距今天数",
            "联系方式",
            "爬取时间",
            "发现来源"
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row, kol in enumerate(kols, 2):
            ws.cell(row=row, column=1, value=kol['channel_name'])
            ws.cell(row=row, column=2, value=kol['channel_url'])
            ws.cell(row=row, column=3, value=kol['subscribers'])
            ws.cell(row=row, column=4, value=f"{kol['ai_ratio']:.1%}")
            ws.cell(row=row, column=5, value=kol['avg_views'])
            ws.cell(row=row, column=6, value=kol['avg_likes'])
            ws.cell(row=row, column=7, value=kol.get('avg_comments', 0))
            ws.cell(row=row, column=8, value=f"{kol['engagement_rate']:.2%}")
            ws.cell(row=row, column=9, value=kol['total_videos'])
            
            # 最后视频日期 - 处理字符串和datetime对象
            if kol.get('last_video_date'):
                if isinstance(kol['last_video_date'], str):
                    # 处理ISO格式时间戳
                    date_str = kol['last_video_date'].split('T')[0] if 'T' in kol['last_video_date'] else kol['last_video_date'][:10]
                    ws.cell(row=row, column=10, value=date_str)
                else:
                    ws.cell(row=row, column=10, value=kol['last_video_date'].strftime('%Y-%m-%d'))
            else:
                ws.cell(row=row, column=10, value="未知")
            
            ws.cell(row=row, column=11, value=kol.get('days_since_last_video') or "未知")
            ws.cell(row=row, column=12, value=kol.get('contact_info') or "")
            
            # 爬取时间 - 处理字符串和datetime对象
            if kol.get('discovered_at'):
                if isinstance(kol['discovered_at'], str):
                    # 处理ISO格式时间戳
                    date_str = kol['discovered_at'].split('T')[0] if 'T' in kol['discovered_at'] else kol['discovered_at'][:10]
                    ws.cell(row=row, column=13, value=date_str)
                else:
                    ws.cell(row=row, column=13, value=kol['discovered_at'].strftime('%Y-%m-%d'))
            else:
                ws.cell(row=row, column=13, value="未知")
            
            ws.cell(row=row, column=14, value=kol.get('discovered_from', ''))
        
        # 调整列宽
        column_widths = [30, 50, 12, 15, 15, 15, 15, 12, 12, 15, 12, 30, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # 保存文件
        output_dir = self.export_config['output_dir']
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.export_config['filename_format'].replace('{timestamp}', timestamp)
        filepath = os.path.join(output_dir, filename)
        
        wb.save(filepath)
        
        logger.info(f"导出完成: {filepath}")
        logger.info(f"共导出 {len(kols)} 个KOL")
        logger.info("=" * 50)
        
        return filepath
